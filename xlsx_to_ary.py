from openpyxl import load_workbook
from tkinter.filedialog import askopenfilename
from tkinter import Tk
import os

EXPORT_FOLDER = 'wiktionary/'
if not os.path.exists(EXPORT_FOLDER):
    os.mkdir(EXPORT_FOLDER)

def get_lang_code(lang_origin):
    translation_table = {'Tamazight':'zgh','Arabic':'ar','French':'fr','English':'en','Spanish':'es','Portuguese':'pt','Turkish':'tr','Latin':'la','Hebrew':'heb','Ancient Greek':'grc'}

    return translation_table[lang_origin]

def translate_part_of_speech(part_of_speech):
    translation_table = {'noun':'سمية','verb':'فيعل','adverb':'حال','adjective':'نعت','pronoun':'ضمير','particle':'حرف','proverb':'متلة ','expressions':'تعبيرات'}

    return translation_table[part_of_speech]

def load_fields(filename):
    wb = load_workbook(filename)
    sheet = wb.active
    dict_list = []
    print(sheet.max_row)
    for i in range(2,sheet.max_row+1):
        if sheet['A'+str(i)].value is not None:
            print(sheet['A'+str(i)].value)
            dict_list.append({'word':sheet['A'+str(i)].value
                             ,'plural':sheet['B'+str(i)].value
                             ,'feminine':sheet['C'+str(i)].value
                             ,'feminine_plural':sheet['D'+str(i)].value
                             ,'lang_origin':sheet['E'+str(i)].value
                             ,'word_origin':sheet['F'+str(i)].value
                             ,'transliteration_origin':sheet['G'+str(i)].value
                             ,'meaning_origin':sheet['H'+str(i)].value
                             ,'certainty_origin':sheet['I'+str(i)].value
                             ,'second_lang_origin':sheet['J'+str(i)].value
                             ,'second_word_origin':sheet['K'+str(i)].value
                             ,'second_transl_origin':sheet['L'+str(i)].value
                             ,'second_meaning_origin':sheet['M'+str(i)].value
                             ,'second_cert_origin':sheet['N'+str(i)].value
                             ,'origins_relationship':sheet['O'+str(i)].value
                             ,'part_of_speech':sheet['P'+str(i)].value
                             ,'pronunciation':sheet['Q'+str(i)].value
                             ,'meaning':sheet['R'+str(i)].value
                             ,'meaning_example':sheet['S'+str(i)].value
                             ,'figurative_meaning':sheet['T'+str(i)].value
                             ,'figurative_example':sheet['U'+str(i)].value
                             ,'synonyms':sheet['V'+str(i)].value
                             ,'antonyms':sheet['W'+str(i)].value
                             ,'dialectal_variations':sheet['X'+str(i)].value
                             ,'related':sheet['Y'+str(i)].value
                             ,'expressions':sheet['Z'+str(i)].value
                             ,'source':sheet['AA'+str(i)].value
                             ,'picture':sheet['AB'+str(i)].value
                             ,'picture_comment':sheet['AC'+str(i)].value})

    return dict_list

def get_origin_code(lang_origin,word_origin,transliteration_origin,meaning_origin
                                                 ,certainty_origin,second_lang_origin,second_word_origin
                                                 ,second_transl_origin,second_meaning_origin,second_cert_origin
                                                 ,origins_relationship):
    if lang_origin is not None:
        if lang_origin is not None:
            code = "=== {{Wt/ary/S|أصل}} ===\n:"
            if certainty_origin == "unsure":
                code+="واقيلا "
            code+=" من {{Wt/ary/étyl|"+get_lang_code(lang_origin)+"|ary|"
            if word_origin is not None:
                code+=word_origin
            if transliteration_origin is not None:
                code+='|'+transliteration_origin
            if meaning_origin is not None:
                code+= "|معنى="+meaning_origin
            code+="}}"
            if second_lang_origin is not None:
                if origins_relationship == 'from':
                    code+=" لّي جات "
                elif origins_relationship == 'or':
                    code+=" ؤلا "
                if second_cert_origin == "unsure":
                    code+="واقيلا"
                code+=" من {{"+"Wt/ary/étyl|{}|ary|".format(get_lang_code(second_lang_origin))
                if second_word_origin is not None:
                    code+=second_word_origin
                if second_transl_origin is not None:
                    code+='|'+second_transl_origin
                
                if second_meaning_origin is not None:
                    code+= "|معنى="+second_meaning_origin
                code+="}}"
        return code

    return ''
    
     

def get_main_section_code(word,part_of_speech,plural,pronunciation,meaning,meaning_example,feminine,feminine_plural):
    code = "=== {"+"{{Wt/ary/S|{}|ary}}".format(translate_part_of_speech(part_of_speech)) +"} ===\n"
    if feminine is None:
        if plural is not None:
            plurals = [w.strip() for w in plural.split('..')]
            if len(plurals) == 1:
                code+="{"+"{{Wt/ary/فرد-جمع1|s=Wt/ary/{}|p=Wt/ary/{}}}".format(word,plurals[0])+"}"
            else:
                code+="{"+"{{Wt/ary/فرد-جمع1|s=Wt/ary/{}|p=Wt/ary/{}|p2=Wt/ary/{}}}".format(word,plurals[0],plurals[1])+"}"
    else:
        if plural is not None:
            plurals = [w.strip() for w in plural.split('..')]
            code+="{"+"{{Wt/ary/فرد-جمع2| د.ف=Wt/ary/{}| د.ج=Wt/ary/{}| ن.ف=Wt/ary/{}| ن.ج=Wt/ary/{}}}".format(word,plurals[0],feminine,feminine_plural)+"}"
        else:
            code+="{"+"{{Wt/ary/فرد-جمع2| د.ف=Wt/ary/{}| ن.ف=Wt/ary/{}}}".format(word,plural[0],feminine,feminine_plural)+"}"
    code+="\n'''{}''' ".format(word)
    pronunciations = [w.strip() for w in pronunciation.split('..')]
    if len(pronunciations) == 1:
        code+="{"+"{{Wt/ary/pron|{}|ary}}".format(pronunciation)+"}" #+meaning
    elif len(pronunciations) > 1:
        code+="{"+"{{Wt/ary/pron|{}|ary}}".format(pronunciations[0])+"}"
        for i in range(1,len(pronunciations)):
            code+=" ؤلا {"+"{{Wt/ary/pron|{}|ary}}".format(pronunciations[i])+"}"
            
    meaning = [m.strip() for m in meaning.split('..')]
    if meaning_example is not None:
        meaning_example = [me.strip() for me in meaning_example.split('..')]
    else:
        meaning_example = []
    for i in range(len(meaning)):
        code+="\n# "+meaning[i]
        if i<len(meaning_example):
            if meaning_example[i] != '':
                code+="\n#*"+meaning_example[i]
    """
    if source is not None:
        code+="<ref>"+source+"</ref>"
    """
    return code

def get_figurative_code(figurative_meaning,figurative_example):
    if figurative_meaning is not None:
        #code = "==== {{Wt/ary/S|مجازيا}} ====\n"
        figurative_meaning = [m.strip() for m in figurative_meaning.split('..')]
        if figurative_example is not None:
            figurative_example = [me.strip() for me in figurative_example.split('..')]
        else:
            figurative_example = []
        for i in range(len(figurative_meaning)):
            code ="\n# '''مجازي: '''"+figurative_meaning[i]
            if i< len(figurative_example):
                if figurative_example[i] != '':
                    code+="\n#*"+figurative_example[i]
        return code

    return ''

def get_synonyms_code(synonyms):
    
    if synonyms is not None:
        synonyms = [w.strip() for w in synonyms.split('.')]
        code = "==== {{Wt/ary/S|مرادفات}} ====\n"
        
        for synonym in synonyms:
            code +="*[[Wt/ary/{}|{}]]\n".format(synonym,synonym)
        return code

    return ''

def get_antonyms_code(antonyms):
    
    if antonyms is not None:
        antonyms = [w.strip() for w in antonyms.split('.')]
        code = "==== {{Wt/ary/S|معاكسات}} ====\n"
        
        for antonym in antonyms:
            code +="*[[Wt/ary/{}|{}]]\n".format(antonym,antonym)
        return code

    return ''

def get_related_code(related):
    if related is not None:
        related = [w.strip() for w in related.split('.')]
        code = "==== {{Wt/ary/S|مشتقات}} ====\n" #{{wt/ary/(}}\n"
        for r in related:
            code +="*[[Wt/ary/{}|{}]]\n".format(r,r)
        #code+="{{wt/ary/(}}"
        return code

    return ''

def get_dialect_var_code(dialectal_variations):
    
    if dialectal_variations is not None:
        dialectal_variations = [w.strip() for w in dialectal_variations.split('.')]
        code = "==== {{Wt/ary/S|شكال لهجاوية}} ====\n"
        
        for dialectal_variation in dialectal_variations:
            code +="*[[|Wt/ary/{}|{}]]\n".format(dialectal_variation,dialectal_variation)
        return code

    return ''

def get_image_code(picture,picture_comment,word):
    
    if picture is not None:
        if picture_comment is None:
            picture_comment = word
            
        code = "[[File:{}|thumb|left|upright|{}".format(picture,picture_comment)+"]]"
        return code
    return ''

def get_expressions_code(expressions):
    if expressions is not None:
        expressions = [ex.strip() for ex in expressions.split('..')]
        code = "==== {{Wt/ary/S|تعبيرات}} ===="
        for ex in expressions:
            code+="\n*[[|Wt/ary/{}|{}]]".format(ex,ex)
        return code

    return ''

def get_source_link_code(source):
    if source is not None:
        sources = [ss.strip() for ss in source.split('..')]
        code = '== {{Wt/ary/S|مصادر}} =='
        for ss in sources:
            code+='\n'+ss
        return code
    return ''

def export_files(full_code,export_filename):
    if not os.path.exists(export_filename):
        full_code = "== {{Wt/ary/langue|ary}} ==\n"+full_code
        with open(export_filename,'w',encoding='utf-8') as f:
            f.write(full_code)
    else:
        with open(export_filename,'a',encoding='utf-8') as f:
            f.write('\n\n'+full_code)

def rnvl(code):
    if code is None or code.strip() == '':
        return ''
    else:
        return code

def get_plural_full_code(word,part_of_speech,plural,plural_pronunciation,feminine,feminine_plural):
    pass

def get_feminine_full_code(word,part_of_speech,plural,feminine_pronunciation,feminine,feminine_plural):
    pass

def get_fem_plural_full_code(word,part_of_speech,plural,fem_plural_pronunciation,feminine,feminine_plural):
    pass
    


if __name__=='__main__':
   
    Tk().withdraw()
    filename = askopenfilename()
    dict_list = load_fields(filename)
    #print(dict_list[0]['word_origin'])
    for i in range(len(dict_list)):
        word               = dict_list[i]['word']
        #print(word)
        part_of_speech     = dict_list[i]['part_of_speech']
        meaning            = dict_list[i]['meaning']
        if word is None or part_of_speech is None or meaning is None:
            continue
        plural             = dict_list[i]['plural']
        feminine           = dict_list[i]['feminine']
        feminine_plural    = dict_list[i]['feminine_plural']
        pronunciation      = dict_list[i]['pronunciation']
        meaning_example    = dict_list[i]['meaning_example']
        figurative_meaning = dict_list[i]['figurative_meaning']
        figurative_example = dict_list[i]['figurative_example']
        source             = dict_list[i]['source']
        lang_origin        = dict_list[i]['lang_origin']
        word_origin        = dict_list[i]['word_origin']
        synonyms           = dict_list[i]['synonyms']
        antonyms           = dict_list[i]['antonyms']
        dialectal_variations = dict_list[i]['dialectal_variations']
        related            = dict_list[i]['related']
        expressions        = dict_list[i]['expressions']
        picture            = dict_list[i]['picture']
        transliteration_origin = dict_list[i]['transliteration_origin']
        meaning_origin     = dict_list[i]['meaning_origin']
        certainty_origin   = dict_list[i]['certainty_origin']
        second_lang_origin = dict_list[i]['second_lang_origin']
        second_word_origin = dict_list[i]['second_word_origin']
        second_transl_origin = dict_list[i]['second_transl_origin']
        second_meaning_origin = dict_list[i]['second_meaning_origin']
        second_cert_origin    = dict_list[i]['second_cert_origin']
        origins_relationship = dict_list[i]['origins_relationship']
        picture_comment      = dict_list[i]['picture_comment']
        

        main_section = get_main_section_code(word,part_of_speech,plural,pronunciation,meaning,meaning_example,feminine,feminine_plural) #done
        ori_section = get_origin_code(lang_origin,word_origin,transliteration_origin,meaning_origin
                                                 ,certainty_origin,second_lang_origin,second_word_origin
                                                 ,second_transl_origin,second_meaning_origin,second_cert_origin
                                                 ,origins_relationship)
        syn_section = get_synonyms_code(synonyms)
        ant_section = get_antonyms_code(antonyms)
        dia_section = get_dialect_var_code(dialectal_variations)
        rel_section = get_related_code(related)
        fig_section = get_figurative_code(figurative_meaning,figurative_example)
        im_section  = get_image_code(picture,picture_comment,word)                  #done
        exp_section = get_expressions_code(expressions)
        src_section = get_source_link_code(source)
        #print(main_section)
        full_code = rnvl('\n'+ori_section+'\n')+rnvl('\n'+im_section+'\n')+'\n'+main_section+rnvl('\n'+fig_section+'\n')
        full_code +=rnvl('\n'+syn_section+'\n')+rnvl('\n'+ant_section+'\n')+rnvl('\n'+dia_section+'\n')+rnvl('\n'+rel_section+'\n')
        full_code +=rnvl('\n'+exp_section)+rnvl('\n'+src_section)
        
        export_files(full_code.strip(),EXPORT_FOLDER+word+'.txt')
        

